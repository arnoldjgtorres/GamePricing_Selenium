import time
from GetSoup import find_prices, get_url_string, enter_buyback_price
import openpyxl
from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import showerror
# from PIL import Image, ImageTk
import os
#import six
import queue as Queue
import threading

start_time = time.time()
print(start_time)
# path = filedialog.askopenfilename(initialdir="/", title="Select file",
#                                filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

'''
path = 'C:\\Users\ArnoldGT\Desktop\ps4_01_08_15.xlsx'
# path = 'C:\\Users\ArnoldGT\Desktop\ns_price_change_8_10_18.xlsx'
# wb = openpyxl.load_workbook('C:\\Users\ArnoldGT\Desktop\ps4_01_08_15.xlsx')
wb = openpyxl.load_workbook(path)
sheet = wb.active
'''

site = "Gamestop "
# game_system = None
# game_system = " Playstation 4 "
# game_system = " Nintendo Switch "
trade_in_str = "Trade In Value "

'''Will read excel sheet for game titles and new/used condition.'''


def read_game(r, sheet):
    item_number = sheet.cell(row=r, column=1).value

    if item_number.find("*") == -1:
        c_flag = 0
        g_title = sheet.cell(row=r, column=9).value
        print("NEW", item_number)
    else:
        c_flag = 1
        g_title = sheet.cell(row=r, column=9).value
        print("USED", item_number)
    return c_flag, g_title


def write_to_sheet(r, cond_flag, price_url_string, game_system, game_title, sheet):
    sheet.cell(row=r, column=10).value = price_url_string
    sell_price = find_prices(price_url_string, cond_flag)
    print("SELL PRICE", sell_price)
    if sell_price == 0:
        original_price = sheet.cell(row=r, column=2).value
        sheet.cell(row=r, column=6).value = original_price
        sell_price = original_price
    else:
        sheet.cell(row=r, column=6).value = sell_price

    if cond_flag == 0:
        price = int(round(float(sell_price)))
        buyback_price = price / 2
        sheet.cell(row=r, column=7).value = buyback_price
        print("NEW BUYBACK", buyback_price)

    if cond_flag == 1:
        buyback_query = site + game_title + " " + game_system + trade_in_str
        credit, cash, bb_str = enter_buyback_price(cond_flag, buyback_query)
        sheet.cell(row=r, column=11).value = bb_str
        sheet.cell(row=r, column=7).value = str(credit) + ", " + str(cash)
        print("USED BUYBACK", credit, cash)


'''Loop through all excel sheet rows to enter sell and buyback prices.'''


def begin_read(system, open_file, to_file):
    game_system = system
    wb_file = open_file.replace('/', '\\')
    wb = openpyxl.load_workbook(wb_file)
    sheet = wb.active
    # print(to_file)
    for r in range(5, sheet.max_row):

        st1 = time.time()
        if sheet.cell(row=r, column=1).value is None:
            break

        cond_flag, game_title = read_game(r, sheet)

        query = str(site) + str(game_title) + " " + str(system)
        #rint(query)

        try:
            price_url_string = get_url_string(query)
        except ValueError:
            print("ERROR PRICE_URL")
            continue
        if not price_url_string.startswith('https://www.gamestop.com'):
            continue

        write_to_sheet(r, cond_flag, price_url_string, game_system, game_title, sheet)

        print("*** %s secs ***" % (time.time() - st1))
        # wb.save("ps4_changed_test2.xlsx")
        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        save_path = desktop + '\\' + to_file + '.xlsx'
        wb.save(save_path)




class Application(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()
        self.create_widgets()


    def create_widgets(self):
        self.file_display = Label(self, text="")
        self.file_display.pack()

        menubar = Menu(self.master)
        self.master.config(menu=menubar)
        file_menu = Menu(menubar, tearoff=0)
        file_menu.add_command(label="Open", command=self.choose_file)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.master.quit)
        menubar.add_cascade(label="File", menu=file_menu)

        system = ""
        file = ""
        tkvar = StringVar(self)
        tkvar.set("None Selected")

        platforms = ['Playstation 4', 'Xbox 1', 'Nintendo Switch']

        self.platform_label = Label(self, text='Choose a platform')
        self.platform_label.pack()
        game_dropdown = OptionMenu(self, tkvar, *platforms, command=self.dropdown)
        game_dropdown.pack()

        self.save_to_label = Label(self, text='Type name for save file')
        self.save_to_label.pack()

        self.save_name = Entry(self)
        self.save_name.pack()

        self.run = Button(self, text="Run Program", fg="red", command=lambda: begin_read(self.system, self.file, self.save_name.get()))
        self.run.pack()

        # self.platform_label.config(text = tkvar.get())

    def dropdown(self, value):
        print(value)
        self.system = value
        return value

    def choose_file(self):
        file = askopenfilename(initialdir="/", title="Select file",
                               filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
        self.file_display.config(text=file)
        self.file = file


    def quit(self):
        sys.exit(0)
'''
'''
root = Tk()
root.geometry("450x250")
root.configure(background='black')

app = Application(master=root)
app.pack(fill=BOTH, expand=YES)
app.mainloop()


#begin_read('Playstation 4', 'C:\\Users\ArnoldGT\Desktop\ps4_01_08_15.xlsx', 'please')
print("--- %s seconds ---" % (time.time() - start_time))
